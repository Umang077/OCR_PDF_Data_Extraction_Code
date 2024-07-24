import PyPDF2
import os
import fitz  # PyMuPDF
import pandas as pd

from openpyxl import load_workbook
def split_pdf_into_pages(pdf_path, output_folder):
    # Create the output folder if it doesn't exist
    os.makedirs(output_folder, exist_ok=True)

    # Open the PDF file
    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        
        # Iterate over each page in the PDF
        for page_number in range(len(reader.pages)):
            # Create a PDF writer object
            writer = PyPDF2.PdfWriter()
            
            # Add the page to the writer
            writer.add_page(reader.pages[page_number])
            
            # Write the page to a new PDF file
            output_pdf_path = os.path.join(output_folder, f"page_{page_number + 1}.pdf")
            with open(output_pdf_path, 'wb') as output_file:
                writer.write(output_file)

def extract_text_and_coords_from_pdf(pdf_path):
    text_coords = []
    with fitz.open(pdf_path) as doc:
        for page_num in range(len(doc)):
            page = doc.load_page(page_num)
            try:
                blocks = page.get_text("dict")["blocks"]
                for b in blocks:
                    for l in b.get("lines", []):
                        for s in l.get("spans", []):
                            text = s.get("text", "")
                            bbox = s.get("bbox", [])
                            text_coords.append((text, bbox))  # Save text and its bounding box coordinates
            except Exception as e:
                print(f"Error processing page {page_num + 1}: {e}")
    return text_coords

def save_to_excel(pages_text, excel_path):

    writer = pd.ExcelWriter(excel_path, engine='xlsxwriter')
    
    for idx, page_text in enumerate(pages_text, start=1):
        sheet_name = f"Sheet{idx}"
       
        text_data = [{'Text': text, 'X1': round(bbox[0], 4), 'Y1': round(bbox[1], 4), 'X2': round(bbox[2], 4), 'Y2': round(bbox[3], 4)} for text, bbox in page_text]

        df = pd.DataFrame(text_data)
        df = df.sort_values(by=['X1', 'Y1', 'X2', 'Y2'])
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    
    writer.save()

if __name__ == "__main__":
    pdf_path = 'Anubhav Shukla Annual Reflections_Revised.pdf'  # Path to your input PDF file
    output_folder = 'output_pages_test'  # Folder to save individual pages
    excel_path = 'output_excel_testing.xlsx'  # Path to output Excel file
    
  
    split_pdf_into_pages(pdf_path, output_folder)

    
    pages_text = []
    
    pdf_files = sorted([file for file in os.listdir(output_folder) if file.endswith(".pdf")], key=lambda x: int(''.join(filter(str.isdigit, x))))
    for file in pdf_files:
        if len(pages_text) >= 6:  
            break
        page_pdf_path = os.path.join(output_folder, file)
        text_coords = extract_text_and_coords_from_pdf(page_pdf_path)
        pages_text.append(text_coords)




    
    save_to_excel(pages_text, excel_path)
wb = load_workbook('final_sheet_1.xlsx')
ws = wb['Sheet1']

for sheet_index in range(1, 7):
   
    df1 = pd.read_excel('master_data_final.xlsx', sheet_name=f'Sheet{sheet_index}')

   
    df2 = pd.read_excel('output_excel_testing.xlsx', sheet_name=f'Sheet{sheet_index}')

    x1_values_1 = df1['X1'].tolist()
    x2_values_1 = df1['X2'].tolist()
    y1_values_1 = df1['Y1'].tolist()
    y2_values_1 = df1['Y2'].tolist()
    x1_values_2 = df2['X1'].tolist()
    x2_values_2 = df2['X2'].tolist()
    y1_values_2 = df2['Y1'].tolist()
    y2_values_2 = df2['Y2'].tolist()

    text_values = df2['Text'].tolist()

   
    start_row = 2  
    while ws[f'DQ{start_row}'].value or ws[f'DM{start_row}'].value is not None:  # Check if DQ column is filled
        start_row += 1

    filtered_text = []
    for base_x1, base_x2, base_y1, base_y2 in zip(x1_values_1, x2_values_1, y1_values_1, y2_values_1):
        var_coordinate = ""
        string = ""
        for read_x1, read_x2, read_y1, read_y2, t_value in zip(x1_values_2, x2_values_2, y1_values_2, y2_values_2,
                                                              text_values):
            base_coordinate = f"{base_x1}_{base_y1}_{base_x2}_{base_y2}"
            var_coordinate = f"{read_x1}_{read_y1}_{read_x2}_{read_y2}"
            if (read_x1 + 4 >= base_x1 and read_x1 - 4 <= base_x1) and (read_y1 + 4 >= base_y1) and (
                    read_x2 <= base_x2) and (read_y2 <= base_y2):
                var_coordinate = base_coordinate
                if var_coordinate == base_coordinate:
                   
                    t_value=str(t_value)
                    string += t_value
        filtered_text.append(string)

    def value(col_offset):
        for col, value in enumerate(filtered_text, start=1):
            ws.cell(row=start_row, column=col + col_offset, value=value)

    col_offsets = [0, 52, 75, 95, 105, 113]

    for col_offset in col_offsets:
        if ws.cell(row=start_row, column=col_offset + 1).value is None:
            value(col_offset)
            break

wb.save('final_sheet_1.xlsx')



